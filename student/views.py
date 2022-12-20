from django.shortcuts import render, HttpResponse, redirect
from django.db import transaction
from django.http import JsonResponse
from student import models

from openpyxl import load_workbook

from utils.res import ResponseData


def add(request):
    class_queryset = models.Classes.objects.all()
    course_queryset = models.Course.objects.all()
    tem_dict = {'class_queryset': class_queryset, 'course_queryset': course_queryset}

    if request.method == 'POST':
        name = request.POST.get('name')
        age = request.POST.get('age')
        classes_id = request.POST.get('classes')
        course_id_list = request.POST.getlist('course')
        if not age:
            age = 18
        try:
            with transaction.atomic():
                stu_obj = models.Student.objects.create(name=name, age=age, classes_id=classes_id)
                course_list = []
                for course_id in course_id_list:
                    course_list.append(models.Student2Course(student=stu_obj, course_id=course_id))
                models.Student2Course.objects.bulk_create(course_list)
        except Exception as e:
            print(e)
            print('服务器错误---student/add!')
        finally:
            return redirect('base:students')

    return render(request, 'student/add.html', tem_dict)


def dels(request):
    if request.is_ajax():
        if request.method == 'POST':
            res_dict = ResponseData()
            current_pk = request.POST.get('current_pk')
            try:
                with transaction.atomic():
                    stu_obj = models.Student.objects.filter(pk=current_pk)
                    student_detail_id = stu_obj.first().student_detail_id
                    if student_detail_id:
                        models.StudentDetail.objects.filter(pk=student_detail_id).delete()
                    stu_obj.delete()
            except Exception as e:
                res_dict.status = 5000
                res_dict.message = '服务器错误---student/dels!'
                print(e)
                print(res_dict.message)
            finally:
                return JsonResponse(res_dict.get_dict)


def edit(request, current_pk):
    stu_obj = models.Student.objects.filter(pk=current_pk).first()
    class_queryset = models.Classes.objects.all()
    course_queryset = models.Course.objects.all()
    course_id_list = [course.pk for course in stu_obj.course.all()]
    tem_dict = {
        'stu_obj': stu_obj,
        'class_queryset': class_queryset,
        'course_queryset': course_queryset,
        'course_id_list': course_id_list
    }
    if request.method == 'POST':
        name = request.POST.get('name')
        age = request.POST.get('age')
        classes_id = request.POST.get('classes')
        course_id_list = request.POST.getlist('course')
        try:
            with transaction.atomic():
                models.Student.objects.filter(pk=current_pk).update(name=name, age=age, classes_id=classes_id)
                models.Student2Course.objects.filter(student=stu_obj).delete()
                course_list = []
                for course_id in course_id_list:
                    course_list.append(models.Student2Course(student=stu_obj, course_id=course_id))
                models.Student2Course.objects.bulk_create(course_list)
        except Exception as e:
            print(e)
            print('服务器错误---student/edit!')
        finally:
            return redirect('base:students')
    return render(request, 'student/edit.html', tem_dict)


def search(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        key_word = request.POST.get('key_word')
        if category == 'name':
            ret_queryset = models.Student.objects.filter(name__contains=key_word)
        elif category == 'classes':
            ret_queryset = models.Student.objects.filter(classes__name__contains=key_word)
        elif category == 'course':
            ret_queryset = models.Student.objects.filter(course__name=key_word)
        else:
            ret_queryset = None
        return render(request, 'student/search.html', {'ret_queryset': ret_queryset, 'key_word': key_word})


def import_student(request):
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            return HttpResponse('请先选择文件')
        wb = load_workbook(import_file)
        sheet = wb.worksheets[0]
        gender_dict = {
            '保密': 0,
            '男': 1,
            '女': 2,
        }
        for cells in sheet.iter_rows():
            name = cells[0].value or None
            if name == '姓名':
                continue
            gender = cells[1].value
            gender_val = gender_dict.get(gender, 0)
            age = cells[2].value or None
            addr = cells[3].value or None
            phone = cells[4].value or None
            classes_name = cells[5].value or None
            try:
                with transaction.atomic():
                    class_obj = models.Classes.objects.filter(name=classes_name).first()
                    student_detail_obj = models.StudentDetail.objects.create(gender=gender_val, addr=addr, phone=phone)
                    models.Student.objects.create(name=name, age=age, classes=class_obj,
                                                  student_detail=student_detail_obj)
            except Exception as e:
                print(e)
                print('服务器错误---student/import_student!')
        return redirect('base:students')
